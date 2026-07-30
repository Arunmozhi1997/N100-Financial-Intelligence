import sqlite3

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DB_PATH = "db/nifty100.db"
PEER_FILE = "data/raw/peer_groups.xlsx"


def load_data():
    """Load peer groups and financial data."""

    conn = sqlite3.connect(DB_PATH)

    peer_groups = pd.read_excel(PEER_FILE)

    financial_ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn,
    )

    analysis = pd.read_sql(
        "SELECT * FROM analysis",
        conn,
    )

    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn,
    )

    conn.close()

    return peer_groups, financial_ratios, analysis, companies


def create_peer_percentiles_table():
    """Create peer_percentiles table if it does not exist."""

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS peer_percentiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id TEXT,
            peer_group_name TEXT,
            metric TEXT,
            value REAL,
            percentile_rank REAL,
            year INTEGER
        )
        """)

    conn.commit()
    conn.close()

    print("\npeer_percentiles table ready.")


def compute_percentiles(peer_groups, financial_ratios):
    """
    Merge peer groups with financial ratios and calculate
    yearly percentile ranks within each peer group.
    """

    # Merge financial data with peer groups
    df = financial_ratios.merge(
        peer_groups,
        on="company_id",
        how="left",
    )

    # Convert year to integer
    df["year"] = pd.to_numeric(
        df["year"],
        errors="coerce",
    )

    df = df.dropna(subset=["year"])

    df["year"] = df["year"].astype(int)

    print("\nMerge Summary")
    print("-" * 40)
    print("Total Rows      :", len(df))
    print("Matched Rows    :", df["peer_group_name"].notna().sum())
    print("Unmatched Rows  :", df["peer_group_name"].isna().sum())

    print("\nCompanies without a Peer Group")
    print("-" * 40)

    print(
        df[df["peer_group_name"].isna()]["company_id"]
        .drop_duplicates()
        .sort_values()
        .tolist()
    )

    metrics = [
        "return_on_equity_pct",
        "net_profit_margin_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "interest_coverage",
        "asset_turnover",
    ]

    for metric in metrics:

        if metric not in df.columns:
            continue

        # Lower Debt-to-Equity is better
        ascending = metric == "debt_to_equity"

        rank = df.groupby(
            [
                "peer_group_name",
                "year",
            ]
        )[metric].rank(
            method="min",
            ascending=ascending,
        )

        total = df.groupby(
            [
                "peer_group_name",
                "year",
            ]
        )[
            metric
        ].transform("count")

        df[f"{metric}_percentile"] = (((total - rank + 1) / total) * 100).round(2)

    print("\nPercentiles Calculated")
    print("-" * 40)

    print(
        df[
            [
                "company_id",
                "year",
                "peer_group_name",
                "return_on_equity_pct",
                "return_on_equity_pct_percentile",
            ]
        ].head(20)
    )

    return df


def save_percentiles(df):
    """Save percentile results into SQLite."""

    conn = sqlite3.connect(DB_PATH)

    cursor = conn.cursor()

    cursor.execute("DELETE FROM peer_percentiles")

    metrics = [
        "return_on_equity_pct",
        "net_profit_margin_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "interest_coverage",
        "asset_turnover",
    ]

    rows = []

    for _, row in df.iterrows():

        if pd.isna(row["peer_group_name"]):
            continue

        for metric in metrics:

            percentile_col = f"{metric}_percentile"

            if percentile_col not in df.columns:
                continue

            value = row.get(metric)

            percentile = row.get(percentile_col)

            if pd.isna(value) or pd.isna(percentile):
                continue

            rows.append(
                (
                    row["company_id"],
                    row["peer_group_name"],
                    metric,
                    float(value),
                    float(percentile),
                    int(row["year"]),
                )
            )

    cursor.executemany(
        """
        INSERT INTO peer_percentiles
        (
            company_id,
            peer_group_name,
            metric,
            value,
            percentile_rank,
            year
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        rows,
    )

    conn.commit()

    print(f"\nInserted {len(rows)} rows into peer_percentiles.")

    conn.close()


def create_peer_comparison(df):
    """Export one Excel sheet per peer group."""

    output_file = "output/peer_comparison.xlsx"

    # Delete old workbook if it already exists
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            print("\nERROR")
            print("----------------------------------------")
            print("Close output/peer_comparison.xlsx and run again.")
            return False

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        peer_groups = sorted(df["peer_group_name"].dropna().unique())

        for group in peer_groups:

            group_df = df[df["peer_group_name"] == group].sort_values(
                by="return_on_equity_pct_percentile",
                ascending=False,
            )

            group_df.to_excel(
                writer,
                sheet_name=group[:31],
                index=False,
            )

    print("\nPeer Comparison Report Saved")
    print(output_file)

    return True


def format_peer_comparison():
    """Apply color formatting and highlight benchmark companies."""

    workbook = load_workbook("output/peer_comparison.xlsx")

    green = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE",
    )

    yellow = PatternFill(
        fill_type="solid",
        start_color="FFF2CC",
        end_color="FFF2CC",
    )

    red = PatternFill(
        fill_type="solid",
        start_color="F4CCCC",
        end_color="F4CCCC",
    )

    gold = PatternFill(
        fill_type="solid",
        start_color="FFD966",
        end_color="FFD966",
    )

    for sheet in workbook.worksheets:

        headers = [cell.value for cell in sheet[1]]

        # -----------------------------
        # Color Percentile Columns
        # -----------------------------
        for col_index, header in enumerate(headers, start=1):

            if header is None:
                continue

            if "percentile" not in str(header).lower():
                continue

            for row in range(2, sheet.max_row + 1):

                cell = sheet.cell(row=row, column=col_index)

                if cell.value is None:
                    continue

                if cell.value >= 75:
                    cell.fill = green

                elif cell.value >= 25:
                    cell.fill = yellow

                else:
                    cell.fill = red

        # -----------------------------
        # Highlight Benchmark Company
        # -----------------------------
        if "is_benchmark" in headers:

            benchmark_col = headers.index("is_benchmark") + 1

            for row in range(2, sheet.max_row + 1):

                value = sheet.cell(
                    row=row,
                    column=benchmark_col,
                ).value

                if str(value).upper() == "TRUE":

                    for col in range(1, sheet.max_column + 1):

                        sheet.cell(
                            row=row,
                            column=col,
                        ).fill = gold

    workbook.save("output/peer_comparison.xlsx")

    print("\nColor formatting applied.")
    print("Benchmark companies highlighted.")


def add_summary_rows():
    """Add median summary row to every peer sheet."""

    workbook = load_workbook("output/peer_comparison.xlsx")

    summary_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAD3",
        end_color="D9EAD3",
    )

    for sheet in workbook.worksheets:

        last_row = sheet.max_row + 1

        sheet.cell(row=last_row, column=1).value = "Median"

        for col in range(2, sheet.max_column + 1):

            values = []

            for row in range(2, last_row):

                value = sheet.cell(row=row, column=col).value

                if isinstance(value, (int, float)):
                    values.append(value)

            if values:

                values.sort()

                median = values[len(values) // 2]

                sheet.cell(
                    row=last_row,
                    column=col,
                ).value = median

        for col in range(1, sheet.max_column + 1):

            sheet.cell(
                row=last_row,
                column=col,
            ).fill = summary_fill

    workbook.save("output/peer_comparison.xlsx")

    print("\nMedian summary rows added.")


if __name__ == "__main__":

    peer_groups, financial_ratios, analysis, companies = load_data()

    print("\nPeer Groups")
    print(peer_groups.shape)

    print("\nFinancial Ratios")
    print(financial_ratios.shape)

    print("\nAnalysis")
    print(analysis.shape)

    print("\nCompanies")
    print(companies.shape)

    merged = compute_percentiles(
        peer_groups,
        financial_ratios,
    )

    create_peer_percentiles_table()

    # Save all percentile values into SQLite
    save_percentiles(merged)

    success = create_peer_comparison(merged)

    if success:
        format_peer_comparison()
        add_summary_rows()

    print("\nSprint 3 Peer Analytics Completed Successfully!")
